"""LOV3 · Adjusted EBITDA Add-Back Schedule generator.

Reads the current Q1 / Q2 / H1 2026 P&L .xlsx files and builds a lender-
facing Add-Back Schedule that bridges Reported EBITDA to Adjusted EBITDA.

Add-back categories (standard SBA/conventional lender convention):

    Definite (uncontested):
        - Owner Draws / Transfers          (equity distribution, not opex)
        - Personal Meals                   (owner personal, not opex)
        - Credit Card Payments             (debt service — below EBITDA)
        - Capital Equipment                (capex — capitalized, not opex)
        - Construction Build-Out           (capex — capitalized, not opex)

    Discretionary (requires owner sign-off):
        - Owner Discretionary Expenses     (Zelle to individuals — mix of
                                            business hosting + personal)

Output:
    - LOV3_HTX_Adjusted_EBITDA_Addback_Schedule.xlsx  (standalone file)
    - Also appended as a sheet to LOV3_HTX_Financial_Statements_SBA.xlsx
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

os.chdir("/Users/maurice_mac/Library/CloudStorage/Dropbox/Developer/Toast_ETL_Workflow")

# ── Line labels to look up in each P&L file ───────────────────────────
LINE_LABELS = {
    "reported_ebitda": ["EBITDA"],
    "revenue":         ["TOTAL OPERATING REVENUE"],
    "owner_draws":     ["Owner Draws"],
    "personal_meals":  ["Personal Meals"],
    "cc_payments":     ["Credit Card Payments"],
    "capital_equip":   ["Capital Equipment"],
    "construction":    ["Construction Build"],
    "owner_disc":      ["Owner Discretionary"],
}

# ── Read the current P&L files ────────────────────────────────────────
def load_period(fname, sheet, ytd_col):
    wb = load_workbook(fname, data_only=True)
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=90, values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        val = row[ytd_col] if len(row) > ytd_col else None
        if val is not None and isinstance(val, (int, float)):
            out[label] = val
    # Map to our short keys
    result = {}
    for key, patterns in LINE_LABELS.items():
        val = 0
        for pattern in patterns:
            for label, v in out.items():
                if pattern.lower() in label.lower() and "margin" not in label.lower() \
                        and "note" not in label.lower() and "total" != label.lower().split()[0]:
                    # Prefer non-TOTAL rows for sub-lines
                    if key == "reported_ebitda":
                        val = v
                    elif key == "revenue":
                        val = v
                    else:
                        val += v
                    break
            if val:
                break
        result[key] = val
    # Reported EBITDA — match exactly "EBITDA" not "EBITDA Margin"
    for label, v in out.items():
        if label.strip() == "EBITDA":
            result["reported_ebitda"] = v
        if "TOTAL OPERATING REVENUE" in label:
            result["revenue"] = v
    return result


q1 = load_period("LOV3_HTX_Q1_2026_PL.xlsx", "Q1 2026 P&L", 4)
q2 = load_period("LOV3_HTX_Q2_2026_PL.xlsx", "Q2 2026 P&L", 4)
h1 = load_period("LOV3_HTX_H1_2026_YTD_PL.xlsx", "H1 2026 YTD P&L", 7)

# ── Styles ─────────────────────────────────────────────────────────────
TITLE = Font(name="Calibri", size=16, bold=True, color="1F3864")
H2 = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LABEL = Font(name="Calibri", size=10)
LABEL_BOLD = Font(name="Calibri", size=10, bold=True)
LABEL_BOLD_GOLD = Font(name="Calibri", size=11, bold=True, color="8A7A3D")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
CURRENCY = "_($* #,##0.00_);_($* (#,##0.00);_($* \"-\"??_);_(@_)"
PERCENT = "0.0%"
FILL_H2 = PatternFill("solid", fgColor="1F3864")
FILL_TOTAL = PatternFill("solid", fgColor="E7E6E6")
FILL_HIGHLIGHT = PatternFill("solid", fgColor="FDF6D7")
FILL_INFO = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(border_style="thin", color="A6A6A6")
MEDIUM = Side(border_style="medium", color="1F3864")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TOTAL_BORDER = Border(top=MEDIUM, bottom=MEDIUM, left=THIN, right=THIN)


def build_sheet(ws):
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    # Title
    ws["A1"] = "LOV3 Restaurant & Lounge LLC"
    ws["A1"].font = TITLE
    ws["A2"] = "Adjusted EBITDA — Add-Back Schedule"
    ws["A2"].font = Font(name="Calibri", size=13, bold=True, color="595959")
    ws["A3"] = "For the Periods Ended March 31 and June 30, 2026"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws["A4"] = f"Prepared: {datetime.now().strftime('%B %d, %Y')}"
    ws["A4"].font = Font(name="Calibri", size=9, italic=True, color="808080")

    # Column headers
    r = 6
    for c, hdr in enumerate(["", "Q1 2026", "Q2 2026", "H1 2026 YTD"], start=1):
        cell = ws.cell(row=r, column=c, value=hdr)
        cell.font = H2
        cell.fill = FILL_H2
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22

    def row(label, values, *, bold=False, fill=None, border=None,
            currency=True, indent=1, top_border_medium=False):
        nonlocal r
        r += 1
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = LABEL_BOLD if bold else LABEL
        c1.alignment = Alignment(vertical="center", indent=indent)
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = LABEL_BOLD if bold else LABEL
            cell.number_format = CURRENCY if currency else PERCENT
            cell.alignment = Alignment(horizontal="right")
            if fill:
                cell.fill = fill
        if fill:
            ws.cell(row=r, column=1).fill = fill
        if border:
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = border
        if top_border_medium:
            for c in range(1, 5):
                cur = ws.cell(row=r, column=c).border or Border()
                ws.cell(row=r, column=c).border = Border(
                    top=MEDIUM, bottom=cur.bottom or THIN,
                    left=cur.left or THIN, right=cur.right or THIN)

    def section(title):
        nonlocal r
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title)
        c.font = H2
        c.fill = FILL_H2
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    # ── STARTING POINT ─────────────────────────────────────────────────
    section("REPORTED EBITDA (per P&L)")
    row("Total Operating Revenue",
        [q1["revenue"], q2["revenue"], h1["revenue"]])
    row("Reported EBITDA",
        [q1["reported_ebitda"], q2["reported_ebitda"], h1["reported_ebitda"]],
        bold=True, fill=FILL_TOTAL)
    row("Reported EBITDA Margin",
        [q1["reported_ebitda"] / q1["revenue"] if q1["revenue"] else 0,
         q2["reported_ebitda"] / q2["revenue"] if q2["revenue"] else 0,
         h1["reported_ebitda"] / h1["revenue"] if h1["revenue"] else 0],
        currency=False)

    # ── DEFINITE ADD-BACKS ─────────────────────────────────────────────
    section("DEFINITE ADD-BACKS (Non-Operating / Below-the-Line)")

    row("Owner Draws / Transfers (equity distribution)",
        [q1["owner_draws"], q2["owner_draws"], h1["owner_draws"]])
    row("Personal Meals (owner spend at other venues)",
        [q1["personal_meals"], q2["personal_meals"], h1["personal_meals"]])
    row("Credit Card Payments (debt service)",
        [q1["cc_payments"], q2["cc_payments"], h1["cc_payments"]])
    row("Capital Equipment (capex — capitalize + depreciate)",
        [q1["capital_equip"], q2["capital_equip"], h1["capital_equip"]])
    row("Construction Build-Out (capex — capitalize)",
        [q1["construction"], q2["construction"], h1["construction"]])

    q1_definite = (q1["owner_draws"] + q1["personal_meals"] + q1["cc_payments"]
                   + q1["capital_equip"] + q1["construction"])
    q2_definite = (q2["owner_draws"] + q2["personal_meals"] + q2["cc_payments"]
                   + q2["capital_equip"] + q2["construction"])
    h1_definite = (h1["owner_draws"] + h1["personal_meals"] + h1["cc_payments"]
                   + h1["capital_equip"] + h1["construction"])
    row("Subtotal — Definite Add-Backs",
        [q1_definite, q2_definite, h1_definite],
        bold=True, fill=FILL_TOTAL, top_border_medium=True)

    # ── ADJUSTED EBITDA ────────────────────────────────────────────────
    section("ADJUSTED EBITDA (Post-Definite Add-Backs)")

    q1_adj = q1["reported_ebitda"] + q1_definite
    q2_adj = q2["reported_ebitda"] + q2_definite
    h1_adj = h1["reported_ebitda"] + h1_definite
    row("Adjusted EBITDA", [q1_adj, q2_adj, h1_adj],
        bold=True, fill=FILL_HIGHLIGHT)
    row("Adjusted EBITDA Margin",
        [q1_adj / q1["revenue"] if q1["revenue"] else 0,
         q2_adj / q2["revenue"] if q2["revenue"] else 0,
         h1_adj / h1["revenue"] if h1["revenue"] else 0],
        bold=True, fill=FILL_HIGHLIGHT, currency=False)

    # ── DISCRETIONARY ADD-BACKS ────────────────────────────────────────
    section("DISCRETIONARY ADD-BACKS (Requires Owner Sign-Off)")

    row("Owner Discretionary Expenses (Zelle to individuals)",
        [q1["owner_disc"], q2["owner_disc"], h1["owner_disc"]])

    # ── FULLY ADJUSTED EBITDA ──────────────────────────────────────────
    section("FULLY ADJUSTED EBITDA")

    q1_full = q1_adj + q1["owner_disc"]
    q2_full = q2_adj + q2["owner_disc"]
    h1_full = h1_adj + h1["owner_disc"]
    row("Fully Adjusted EBITDA",
        [q1_full, q2_full, h1_full],
        bold=True, fill=FILL_HIGHLIGHT, top_border_medium=True)
    row("Fully Adjusted EBITDA Margin",
        [q1_full / q1["revenue"] if q1["revenue"] else 0,
         q2_full / q2["revenue"] if q2["revenue"] else 0,
         h1_full / h1["revenue"] if h1["revenue"] else 0],
        bold=True, fill=FILL_HIGHLIGHT, currency=False)

    # ── FOOTNOTES ──────────────────────────────────────────────────────
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r+7, end_column=4)
    notes = ws.cell(row=r, column=1, value=(
        "METHODOLOGY NOTES:\n"
        "  • Owner Draws / Transfers = equity distributions to LLC members. Balance sheet item, not P&L expense.\n"
        "  • Personal Meals = owner spend at other restaurants (Clarkwood HT, etc.). Not operating expense.\n"
        "  • Credit Card Payments = debt service (principal + interest). EBITDA excludes Interest by definition.\n"
        "  • Capital Equipment + Construction Build-Out = capital expenditures. EBITDA excludes Depreciation.\n"
        "  • Owner Discretionary = Zelle payments to individuals (OLAJIDE OLUKOGA, Lou, Chance, Big Tone, Mom).\n"
        "    Mix of business hosting and personal. Requires owner sign-off before treating as add-back.\n"
        "  • Add-back schedule follows standard SBA / conventional lender presentation convention.\n"
    ))
    notes.font = Font(name="Calibri", size=9, italic=True, color="595959")
    notes.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    notes.fill = FILL_INFO
    ws.row_dimensions[r].height = 100

    return q1_adj, q2_adj, h1_adj, q1_full, q2_full, h1_full


# ── Standalone file ───────────────────────────────────────────────────
wb1 = Workbook()
wb1.remove(wb1.active)
ws1 = wb1.create_sheet("Adjusted EBITDA Add-Back")
q1_adj, q2_adj, h1_adj, q1_full, q2_full, h1_full = build_sheet(ws1)
wb1.save("LOV3_HTX_Adjusted_EBITDA_Addback_Schedule.xlsx")
print("Saved: LOV3_HTX_Adjusted_EBITDA_Addback_Schedule.xlsx")

# ── Also append into the master workbook so the lender gets one file ──
main_wb = load_workbook("LOV3_HTX_Financial_Statements_SBA.xlsx")
if "Adjusted EBITDA Add-Back" in main_wb.sheetnames:
    del main_wb["Adjusted EBITDA Add-Back"]
ws2 = main_wb.create_sheet("Adjusted EBITDA Add-Back")
build_sheet(ws2)
main_wb.save("LOV3_HTX_Financial_Statements_SBA.xlsx")
print("Appended sheet to: LOV3_HTX_Financial_Statements_SBA.xlsx")

print()
print("=== Summary ===")
print(f"{'Metric':40s} {'Q1':>12s} {'Q2':>12s} {'H1 YTD':>12s}")
print("-" * 82)
print(f"{'Reported EBITDA':40s} ${q1['reported_ebitda']:>11,.0f} ${q2['reported_ebitda']:>11,.0f} ${h1['reported_ebitda']:>11,.0f}")
print(f"{'  Reported Margin':40s} {q1['reported_ebitda']/q1['revenue']*100:>11.1f}% {q2['reported_ebitda']/q2['revenue']*100:>11.1f}% {h1['reported_ebitda']/h1['revenue']*100:>11.1f}%")
print(f"{'Adjusted EBITDA (definite)':40s} ${q1_adj:>11,.0f} ${q2_adj:>11,.0f} ${h1_adj:>11,.0f}")
print(f"{'  Adjusted Margin':40s} {q1_adj/q1['revenue']*100:>11.1f}% {q2_adj/q2['revenue']*100:>11.1f}% {h1_adj/h1['revenue']*100:>11.1f}%")
print(f"{'Fully Adjusted EBITDA':40s} ${q1_full:>11,.0f} ${q2_full:>11,.0f} ${h1_full:>11,.0f}")
print(f"{'  Fully Adjusted Margin':40s} {q1_full/q1['revenue']*100:>11.1f}% {q2_full/q2['revenue']*100:>11.1f}% {h1_full/h1['revenue']*100:>11.1f}%")
