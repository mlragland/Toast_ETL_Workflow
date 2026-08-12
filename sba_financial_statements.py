"""
SBA Loan Financial Statement Generator — LOV3|HTX

Generates a professional Excel workbook with:
  Sheet 1: 2025 Year-End P&L (Jan–Dec monthly + YTD)
  Sheet 2: Feb 2026 Interim P&L (Jan–Feb monthly + YTD)

Usage:
  pip install openpyxl google-cloud-bigquery
  python sba_financial_statements.py
"""

import calendar
import logging
from datetime import date
from typing import Dict, List, Tuple

from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    DATASET_ID,
    GRAT_PASSTHROUGH_PCT,
    GRAT_RETAIN_PCT,
    PROJECT_ID,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
log = logging.getLogger(__name__)

# ── Styles ───────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=11, bold=False, italic=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
HEADER_FONT = Font(name="Calibri", size=10, bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
CURRENCY_FMT = '$#,##0'
PCT_FMT = '0.0%'
THIN_TOP = Border(top=Side(style="thin"))
THIN_BOTTOM = Border(bottom=Side(style="thin"))
DOUBLE_BOTTOM = Border(bottom=Side(style="double"), top=Side(style="thin"))
RED_FONT = Font(name="Calibri", size=10, color="FF0000")

# ── P&L Line Item Definitions ────────────────────────────────────────────────
# Each tuple: (label, key, indent, line_type)
# line_type: "section", "item", "total", "subtotal", "blank", "net_income"

PNL_STRUCTURE = [
    ("REVENUE", None, 0, "section"),
    ("Food Sales", "food_rev", 1, "item"),
    ("Liquor / Beverage Sales", "liquor_rev", 1, "item"),
    ("Hookah Sales", "hookah_rev", 1, "item"),
    ("Other Revenue", "other_rev", 1, "item"),
    ("Gratuity Collected (20% auto)", "gratuity", 1, "item"),
    ("Tips Collected (voluntary)", "tips", 1, "item"),
    ("Cash Sales (undeposited)", "cash_undeposited", 1, "item"),
    ("Sales Tax Collected", "sales_tax", 1, "item"),
    ("TOTAL OPERATING REVENUE", "total_revenue", 0, "total"),
    ("", None, 0, "blank"),
    ("COST OF GOODS SOLD", None, 0, "section"),
    ("Food COGS", "food_cogs", 1, "item"),
    ("Liquor / Beverage COGS", "liquor_cogs", 1, "item"),
    ("Supplies & Smallwares", "supplies_cogs", 1, "item"),
    ("TOTAL COGS", "total_cogs", 0, "total"),
    ("", None, 0, "blank"),
    ("GROSS PROFIT", "gross_profit", 0, "total"),
    ("  Gross Margin %", "gross_margin_pct", 0, "pct_line"),
    ("", None, 0, "blank"),
    ("LABOR", None, 0, "section"),
    ("Payroll & Wages", "labor_gross", 1, "item"),
    ("  Labor % of Revenue", "labor_pct", 0, "pct_line"),
    ("", None, 0, "blank"),
    ("MARKETING & ENTERTAINMENT", None, 0, "section"),
    ("Entertainment", "mkt_entertainment", 1, "item"),
    ("Promoter Payout", "mkt_promoter", 1, "item"),
    ("PMG / Artist Booking", "mkt_artist", 1, "item"),
    ("Social Media", "mkt_social", 1, "item"),
    ("Flyers & Print", "mkt_flyers", 1, "item"),
    ("Event Expense", "mkt_event", 1, "item"),
    ("Pay-Per-View", "mkt_ppv", 1, "item"),
    ("TOTAL MARKETING", "total_marketing", 0, "total"),
    ("", None, 0, "blank"),
    ("OPERATING EXPENSES", None, 0, "section"),
    ("Rent & CAM", "opex_rent", 1, "item"),
    ("Taxes", "opex_taxes", 1, "item"),
    ("Security", "opex_security", 1, "item"),
    ("Insurance", "opex_insurance", 1, "item"),
    ("Bussers & Cleaners", "opex_bussers", 1, "item"),
    ("Contract Labor", "opex_contract_labor", 1, "item"),
    ("Janitorial Services", "opex_cleaning", 1, "item"),
    ("Utilities", "opex_utilities", 1, "item"),
    ("POS & Technology Fees", "opex_pos_tech", 1, "item"),
    ("Software & Subscriptions", "opex_software", 1, "item"),
    ("Phone & Internet", "opex_phone", 1, "item"),
    ("Professional Services", "opex_professional", 1, "item"),
    ("Permits & Licenses", "opex_permits", 1, "item"),
    ("Bank Fees", "opex_bank_fees", 1, "item"),
    ("Penalties & Fees", "opex_penalties", 1, "item"),
    ("Admin & Office", "opex_admin", 1, "item"),
    ("Lighting & Sound", "opex_lighting", 1, "item"),
    ("Other / Uncategorized", "opex_other", 1, "item"),
    ("TOTAL OPERATING EXPENSES", "total_opex", 0, "total"),
    ("", None, 0, "blank"),
    ("G&A / CORPORATE", None, 0, "section"),
    ("Owner Draws / Transfers", "ga_owner_draws", 1, "item"),
    ("Owner Discretionary", "ga_discretionary", 1, "item"),
    ("Personal Meals", "ga_meals", 1, "item"),
    ("Transportation", "ga_transportation", 1, "item"),
    ("Travel & Lodging", "ga_travel", 1, "item"),
    ("Credit Card Payments", "ga_credit_card", 1, "item"),
    ("Competitive Research", "ga_competitive", 1, "item"),
    ("Other G&A", "ga_other", 1, "item"),
    ("TOTAL G&A", "total_ga", 0, "total"),
    ("", None, 0, "blank"),
    ("FACILITY & TENANT IMPROVEMENTS", None, 0, "section"),
    ("Construction Build-Out", "cap_construction", 1, "item"),
    ("Capital Equipment", "cap_equipment", 1, "item"),
    ("Repairs & Maintenance", "cap_repairs", 1, "item"),
    ("TOTAL FACILITY & TI", "total_capex", 0, "total"),
    ("", None, 0, "blank"),
    ("TOTAL EXPENSES", "total_all_expenses", 0, "total"),
    ("", None, 0, "blank"),
    ("EBITDA", "ebitda", 0, "net_income"),
    ("  EBITDA Margin %", "ebitda_pct", 0, "pct_line"),
    ("", None, 0, "blank"),
    ("Note: Revenue includes tip & gratuity pass-through to staff", "pass_through_memo", 0, "memo"),
]


# ── BigQuery Queries ─────────────────────────────────────────────────────────

def _make_date_config(start: str, end: str) -> bigquery.QueryJobConfig:
    """Parameterized query config for Toast tables (DATE type)."""
    return bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("start_date", "DATE", start),
        bigquery.ScalarQueryParameter("end_date", "DATE", end),
    ])


def _make_string_config(start: str, end: str) -> bigquery.QueryJobConfig:
    """Parameterized query config for BankTransactions (STRING dates)."""
    return bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("start_date", "STRING", start),
        bigquery.ScalarQueryParameter("end_date", "STRING", end),
    ])


def query_monthly_revenue(client: bigquery.Client, start: str, end: str) -> Dict[str, Dict]:
    """Monthly net_sales, tips, gratuity from OrderDetails_raw."""
    q = f"""
    SELECT
        FORMAT_DATE('%Y-%m', processing_date) AS month,
        COALESCE(SUM(amount), 0) AS net_sales,
        COALESCE(SUM(tip), 0) AS tips,
        COALESCE(SUM(gratuity), 0) AS gratuity,
        COUNT(DISTINCT order_id) AS order_count
    FROM `{PROJECT_ID}.{DATASET_ID}.OrderDetails_raw`
    WHERE processing_date BETWEEN @start_date AND @end_date
        AND (voided IS NULL OR voided = 'false')
    GROUP BY month ORDER BY month
    """
    rows = client.query(q, job_config=_make_date_config(start, end)).result()
    return {
        r.month: {
            "net_sales": float(r.net_sales or 0),
            "tips": float(r.tips or 0),
            "gratuity": float(r.gratuity or 0),
            "order_count": int(r.order_count or 0),
        }
        for r in rows
    }


def query_revenue_by_category(client: bigquery.Client, start: str, end: str) -> Dict[str, Dict]:
    """Monthly food vs liquor revenue from ItemSelectionDetails_raw."""
    q = f"""
    SELECT
        FORMAT_DATE('%Y-%m', processing_date) AS month,
        COALESCE(SUM(CASE WHEN sales_category = 'Food' THEN CAST(net_price AS FLOAT64) ELSE 0 END), 0) AS food_rev,
        COALESCE(SUM(CASE WHEN sales_category = 'Liquor' THEN CAST(net_price AS FLOAT64) ELSE 0 END), 0) AS liquor_rev
    FROM `{PROJECT_ID}.{DATASET_ID}.ItemSelectionDetails_raw`
    WHERE processing_date BETWEEN @start_date AND @end_date
        AND (voided IS NULL OR voided = 'false')
    GROUP BY month ORDER BY month
    """
    rows = client.query(q, job_config=_make_date_config(start, end)).result()
    return {
        r.month: {
            "food_rev": float(r.food_rev or 0),
            "liquor_rev": float(r.liquor_rev or 0),
        }
        for r in rows
    }


def query_hookah_revenue_bank(client: bigquery.Client, start: str, end: str) -> Dict[str, float]:
    """Monthly hookah revenue from bank deposits (Predictive Insights, May 2025+)."""
    q = f"""
    SELECT
        FORMAT_DATE('%Y-%m', transaction_date) AS month,
        COALESCE(SUM(amount), 0) AS hookah_rev
    FROM `{PROJECT_ID}.{DATASET_ID}.BankTransactions_raw`
    WHERE transaction_date BETWEEN @start_date AND @end_date
        AND LOWER(category) LIKE '%hookah sales%'
        AND amount > 0
    GROUP BY month ORDER BY month
    """
    rows = client.query(q, job_config=_make_date_config(start, end)).result()
    return {r.month: float(r.hookah_rev or 0) for r in rows}


def query_hookah_revenue_pos(client: bigquery.Client, start: str, end: str) -> Dict[str, float]:
    """Monthly hookah revenue from Toast POS (in-house, through Mar 2025)."""
    q = f"""
    SELECT
        FORMAT_DATE('%Y-%m', processing_date) AS month,
        COALESCE(SUM(CAST(net_price AS FLOAT64)), 0) AS hookah_rev
    FROM `{PROJECT_ID}.{DATASET_ID}.ItemSelectionDetails_raw`
    WHERE processing_date BETWEEN @start_date AND @end_date
        AND sales_category = 'Hookah'
        AND (voided IS NULL OR voided = 'false')
    GROUP BY month ORDER BY month
    """
    rows = client.query(q, job_config=_make_date_config(start, end)).result()
    return {r.month: float(r.hookah_rev or 0) for r in rows}


# Predictive Insights $20K from Jan 2024 reclassed to Apr 2025 hookah revenue
HOOKAH_RECLASS = {
    "2025-04": 20_000.00,
    "2025-12": 15_000.00,
    "2026-03": 16_400.00,
    # Q2 2026 hookah payments (Predictive Insights) — received off-bank
    # channels (check/Zelle) per operator; reclassed onto the P&L to
    # capture full hookah revenue for the SBA package.
    "2026-04": 16_860.00,
    "2026-05": 20_785.00,
    "2026-06": 19_730.00,
}


def query_expenses_by_category(client: bigquery.Client, start: str, end: str) -> Dict[str, Dict[str, float]]:
    """Monthly expenses grouped by bank category, with a normalization pass
    that maps Plaid's ALL_CAPS_UNDERSCORE taxonomy AND vendor-based fallbacks
    onto the LOV3 numbered taxonomy the keyword mapping expects.

    The Plaid taxonomy started appearing in BankTransactions_raw when the
    Plaid pipeline went live on 2026-07-30 (backfilled Jan onward). The
    original LOV3 numbered taxonomy ("1. Revenue Recognition/…", "5. Operating
    Expenses (OPEX)/Rent / CAM / …") continued being used for pre-existing
    rows. Post-2026-07-30 rows in Q1 also carry Plaid categories.

    Without this normalization, Marketing dropped 63% and G&A dropped 48%
    quarter-over-quarter — not because spend fell, but because the Plaid-
    categorized rows fell into "Uncategorized" and got dropped.
    """
    q = f"""
    WITH raw AS (
      SELECT
          FORMAT_DATE('%Y-%m', transaction_date) AS month,
          transaction_date,
          category,
          COALESCE(vendor_normalized, '') AS vendor,
          COALESCE(description, '') AS description,
          abs_amount
      FROM `{PROJECT_ID}.{DATASET_ID}.BankTransactions_raw`
      WHERE transaction_date BETWEEN @start_date AND @end_date
          AND transaction_type = 'debit'
    ),
    normalized AS (
      SELECT
        month,
        CASE
          -- 0) Vendor-based overrides — apply BEFORE Plaid category mapping
          --    because Plaid mislabels these specific vendors.
          --
          --    JCS Bar Supplies → bar mixers/glassware/supplies (Plaid tags
          --    them as HOME_IMPROVEMENT_HARDWARE, but they're COGS).
          WHEN REGEXP_CONTAINS(LOWER(vendor), r'jcs bar supplies')
               OR REGEXP_CONTAINS(LOWER(description), r'jcs bar supplies')
               THEN '2. Cost of Goods Sold/Supplies & Equipment'
          --    Kraftsmen — bread/bakery supplier (Plaid tags as HOME_IMPROVEMENT).
          WHEN REGEXP_CONTAINS(LOWER(vendor), r'kraftsmen')
               OR REGEXP_CONTAINS(LOWER(description), r'kraftsmen')
               THEN '2. Cost of Goods Sold/Food COGS'
          --    Kitchens — kitchen equipment supplier (Plaid tags as
          --    GENERAL_SERVICES_OTHER_GENERAL_SERVICES, but it's Facility CapEx).
          WHEN REGEXP_CONTAINS(LOWER(vendor), r'^kitchens\b|kitchens\s|the kitchens')
               THEN '7. Facility & Tenant Improvements/Capital Equipment Expense'
          --    Wave — event ticketing platform (Plaid tags as
          --    GENERAL_SERVICES_OTHER_GENERAL_SERVICES, but it's Marketing/Event).
          WHEN REGEXP_CONTAINS(LOWER(vendor), r'^wave\b|^wave$|wave.*unforgett')
               OR REGEXP_CONTAINS(LOWER(description), r'wave.*unforgett|wave-\*unforgett')
               THEN '4. Marketing & Promotions Expense/Event Expense'
          --    Cintas — uniforms + laundry (Plaid tags as GENERAL_SERVICES,
          --    routes to Professional Services which is wrong).
          WHEN REGEXP_CONTAINS(LOWER(vendor), r'cintas')
               OR REGEXP_CONTAINS(LOWER(description), r'cintas')
               THEN '5. Operating Expenses (OPEX)/Admin & Office'
          --    Scentair — scent machines (Plaid: HOME_IMPROVEMENT). Small
          --    subscription-style OPEX, not Facility CapEx.
          WHEN REGEXP_CONTAINS(LOWER(vendor), r'scentair')
               THEN '5. Operating Expenses (OPEX)/Software & Subscriptions'
          --    BAED Corporation — LOV3's CPA (as of 2025+) who handles all
          --    tax payments on behalf of the business. Transfers to BAED's
          --    account (CHK 4115) fund federal / franchise / property / TABC
          --    tax obligations. Prior categorization as "Professional
          --    Services" was technically correct (BAED is a CPA) but the
          --    economic substance is tax remittance.
          --    NOTE: 2024 BAED payments were payroll (BAED was the payroll
          --    processor pre-Choice Employer). Restrict override to 2025+.
          WHEN (REGEXP_CONTAINS(LOWER(vendor), r'baed')
                OR REGEXP_CONTAINS(LOWER(description), r'baed corporation'))
               AND transaction_date >= '2025-01-01'
               THEN '5. Operating Expenses (OPEX)/Taxes'

          -- 1) Keep the LOV3 numbered taxonomy as-is (highest signal)
          WHEN REGEXP_CONTAINS(category, r'^[0-9]\.') THEN category

          -- 2) Plaid taxonomy → LOV3 destination
          WHEN category = 'FOOD_AND_DRINK_BEER_WINE_AND_LIQUOR'
               THEN '2. Cost of Goods Sold/Liquor / Beverage COGS'
          WHEN category IN ('FOOD_AND_DRINK_GROCERIES')
               THEN '2. Cost of Goods Sold/Food COGS'
          WHEN category IN ('GENERAL_MERCHANDISE_ONLINE_MARKETPLACES',
                            'GENERAL_MERCHANDISE_OTHER_GENERAL_MERCHANDISE',
                            'GENERAL_MERCHANDISE_SUPERSTORES',
                            'GENERAL_MERCHANDISE_CONVENIENCE_STORES',
                            'GENERAL_MERCHANDISE_DEPARTMENT_STORES')
               THEN '2. Cost of Goods Sold/Supplies & Equipment'
          WHEN category = 'HOME_IMPROVEMENT_HARDWARE'
               THEN '5. Operating Expenses (OPEX)/Repairs & Maintenance'
          WHEN category IN ('HOME_IMPROVEMENT_REPAIR_AND_MAINTENANCE',
                            'HOME_IMPROVEMENT_OTHER_HOME_IMPROVEMENT')
               THEN '5. Operating Expenses (OPEX)/Repairs & Maintenance'
          WHEN category = 'HOME_IMPROVEMENT_SECURITY'
               THEN '5. Operating Expenses (OPEX)/Security'
          WHEN category = 'RENT_AND_UTILITIES_RENT'
               THEN '5. Operating Expenses (OPEX)/Rent / CAM / Property Taxes'
          WHEN category IN ('RENT_AND_UTILITIES_GAS_AND_ELECTRICITY',
                            'RENT_AND_UTILITIES_WATER',
                            'RENT_AND_UTILITIES_SEWAGE_AND_WASTE_MANAGEMENT')
               THEN '5. Operating Expenses (OPEX)/Utility -- Electric & Gas'
          WHEN category = 'RENT_AND_UTILITIES_INTERNET_AND_CABLE'
               THEN '5. Operating Expenses (OPEX)/Phone & Internet'
          WHEN category = 'RENT_AND_UTILITIES_TELEPHONE'
               THEN '5. Operating Expenses (OPEX)/Phone & Internet'
          WHEN category = 'GENERAL_SERVICES_INSURANCE'
               THEN '5. Operating Expenses (OPEX)/Insurance'
          WHEN category = 'GENERAL_SERVICES_OTHER_GENERAL_SERVICES'
               THEN '5. Operating Expenses (OPEX)/Professional Services (Consulting, Accounting, Legal)'
          WHEN category = 'GENERAL_SERVICES_STORAGE'
               THEN '5. Operating Expenses (OPEX)/Repairs & Maintenance'
          WHEN category = 'GENERAL_SERVICES_ACCOUNTING_AND_FINANCIAL_PLANNING'
               THEN '5. Operating Expenses (OPEX)/Professional Services (Consulting, Accounting, Legal)'
          WHEN category = 'GENERAL_SERVICES_CONSULTING_AND_LEGAL'
               THEN '5. Operating Expenses (OPEX)/Professional Services (Consulting, Accounting, Legal)'
          WHEN category = 'GENERAL_SERVICES_ADVERTISING_AND_MARKETING'
               THEN '4. Marketing & Promotions Expense/Social Media Marketing'
          WHEN category = 'GENERAL_SERVICES_AUTOMOTIVE'
               THEN '6. General & Administrative / Corporate/Transportation'
          WHEN category = 'GOVERNMENT_AND_NON_PROFIT_TAX_PAYMENT'
               THEN '5. Operating Expenses (OPEX)/Taxes'
          WHEN category IN ('GOVERNMENT_AND_NON_PROFIT_OTHER_GOVERNMENT_AND_NON_PROFIT',
                            'GOVERNMENT_AND_NON_PROFIT_GOVERNMENT_DEPARTMENTS_AND_AGENCIES')
               THEN '5. Operating Expenses (OPEX)/Permits & Licenses'
          WHEN category = 'LOAN_PAYMENTS_CREDIT_CARD_PAYMENT'
               THEN '6. General & Administrative / Corporate/Credit Card Payments'
          WHEN category IN ('LOAN_PAYMENTS_MORTGAGE_PAYMENT',
                            'LOAN_PAYMENTS_STUDENT_LOAN_PAYMENT',
                            'LOAN_PAYMENTS_CAR_PAYMENT',
                            'LOAN_PAYMENTS_PERSONAL_LOAN_PAYMENT')
               THEN '6. General & Administrative / Corporate/Owner Discretionary Expenses'
          WHEN category IN ('BANK_FEES_ATM_FEES', 'BANK_FEES_FOREIGN_TRANSACTION_FEES',
                            'BANK_FEES_INSUFFICIENT_FUNDS', 'BANK_FEES_INTEREST_CHARGE',
                            'BANK_FEES_OVERDRAFT_FEES', 'BANK_FEES_OTHER_BANK_FEES')
               THEN '5. Operating Expenses (OPEX)/Bank Fees'
          WHEN category IN ('ENTERTAINMENT_OTHER_ENTERTAINMENT',
                            'ENTERTAINMENT_TV_AND_MOVIES',
                            'ENTERTAINMENT_MUSIC_AND_AUDIO',
                            'ENTERTAINMENT_VIDEO_GAMES')
               THEN '4. Marketing & Promotions Expense/Entertainment'
          WHEN category IN ('ENTERTAINMENT_CASINOS_AND_GAMBLING',
                            'ENTERTAINMENT_SPORTING_EVENTS_AMUSEMENT_PARKS_AND_MUSEUMS')
               THEN '6. General & Administrative / Corporate/Owner Discretionary Expenses'
          WHEN category IN ('FOOD_AND_DRINK_RESTAURANT', 'FOOD_AND_DRINK_FAST_FOOD',
                            'FOOD_AND_DRINK_COFFEE', 'FOOD_AND_DRINK_OTHER_FOOD_AND_DRINK')
               THEN '6. General & Administrative / Corporate/Personal Meals'
          WHEN category IN ('TRAVEL_FLIGHTS', 'TRAVEL_LODGING', 'TRAVEL_TAXIS_AND_RIDE_SHARES',
                            'TRAVEL_GAS', 'TRAVEL_PUBLIC_TRANSIT', 'TRAVEL_OTHER_TRAVEL',
                            'TRAVEL_PARKING', 'TRAVEL_RENTAL_CARS_AND_TAXIS')
               THEN '6. General & Administrative / Corporate/Owner Travel'
          -- CHOICE EMPLOYER SOLUTIONS wires are payroll (LOV3's PEO). Route to Labor.
          WHEN category IN ('TRANSFER_OUT_WIRE', 'TRANSFER_OUT_ACCOUNT_TRANSFER')
               AND (REGEXP_CONTAINS(UPPER(description), r'CHOICE EMPLOYER')
                    OR REGEXP_CONTAINS(UPPER(vendor), r'CHOICE EMPLOYER'))
               THEN '3. Labor Cost (Includes Grat + Tips)/Employee Payroll (FOH, BOH, Salaries & Taxes)'
          WHEN category IN ('TRANSFER_OUT_WIRE', 'TRANSFER_OUT_ACCOUNT_TRANSFER',
                            'TRANSFER_OUT_INVESTMENT_AND_RETIREMENT_FUNDS',
                            'TRANSFER_OUT_SAVINGS', 'TRANSFER_OUT_WITHDRAWAL',
                            'TRANSFER_OUT_OTHER_TRANSFER_OUT')
               THEN '6. General & Administrative / Corporate/Owner Draws (Cash Withdrawals from ATMs & CC)'
          WHEN category IN ('PERSONAL_CARE_HAIR_AND_BEAUTY', 'PERSONAL_CARE_GYMS_AND_FITNESS_CENTERS',
                            'PERSONAL_CARE_OTHER_PERSONAL_CARE', 'PERSONAL_CARE_LAUNDRY_AND_DRY_CLEANING')
               THEN '6. General & Administrative / Corporate/Owner Discretionary Expenses'
          WHEN category IN ('MEDICAL_DENTAL_CARE', 'MEDICAL_EYE_CARE', 'MEDICAL_NURSING_CARE',
                            'MEDICAL_PHARMACIES_AND_SUPPLEMENTS', 'MEDICAL_PRIMARY_CARE',
                            'MEDICAL_VETERINARY_SERVICES', 'MEDICAL_OTHER_MEDICAL')
               THEN '6. General & Administrative / Corporate/Owner Discretionary Expenses'

          -- 3) Uncategorized fallback → route by vendor / description signal
          --    (Uncategorized was a $245K bucket in H1 before this pass.)
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND (REGEXP_CONTAINS(LOWER(vendor), r'kelvin boj')
                    OR REGEXP_CONTAINS(LOWER(description), r'kelvin boj|promoter'))
               THEN '4. Marketing & Promotions Expense/Promoter Payout'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description),
                    r'\\bdj\\b|djeric|dj eric|d\\.j\\.|artist booking')
               THEN '4. Marketing & Promotions Expense/PMG Artist Booking'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(vendor),
                    r'seaton|jankowski|charlie pena|zo frost')
               THEN '5. Operating Expenses (OPEX)/Contract Labor'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description),
                    r'food and beverag|sysco|rndc|southern glazer|lonestar fruit|spec.s family')
               THEN '2. Cost of Goods Sold/Food COGS'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description),
                    r'miami gardens|las vegas|k kel next to|lincoln capital|bowlmor')
               THEN '6. General & Administrative / Corporate/Owner Discretionary Expenses'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description), r'vendome|owner travel')
               THEN '6. General & Administrative / Corporate/Owner Travel'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description), r'tst\* ?lov3|lov3 restaur')
               THEN '6. General & Administrative / Corporate/Personal Meals'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description), r'sba loan|loan payment|loan svc')
               THEN '6. General & Administrative / Corporate/Credit Card Payments'
          -- Event vendors / performers previously left in Uncategorized bucket
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description),
                    r'wave.*unforgett|unforgettable|emerging 100')
               THEN '4. Marketing & Promotions Expense/Event Expense'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND (REGEXP_CONTAINS(LOWER(vendor), r'aren andoun|omari joseph')
                    OR REGEXP_CONTAINS(LOWER(description), r'aren andoun|omari joseph'))
               THEN '4. Marketing & Promotions Expense/PMG Artist Booking'
          WHEN (category IS NULL OR category = '' OR LOWER(category) = 'uncategorized')
               AND REGEXP_CONTAINS(LOWER(description), r'^check ')
               THEN '5. Operating Expenses (OPEX)/Bussers & Cleaners'

          -- 4) Otherwise keep as-is (will still fall into Other/Uncategorized bucket)
          ELSE COALESCE(NULLIF(category, ''), 'Uncategorized')
        END AS category,
        abs_amount
      FROM raw
    )
    SELECT month, category, ROUND(SUM(abs_amount), 2) AS total
    FROM normalized
    GROUP BY month, category
    ORDER BY month, total DESC
    """
    rows = client.query(q, job_config=_make_date_config(start, end)).result()
    result: Dict[str, Dict[str, float]] = {}
    for r in rows:
        if r.month not in result:
            result[r.month] = {}
        result[r.month][r.category] = float(r.total or 0)
    return result


def query_sales_tax(client: bigquery.Client, start: str, end: str) -> Dict[str, float]:
    """Monthly sales tax collected from OrderDetails_raw."""
    q = f"""
    SELECT
        FORMAT_DATE('%Y-%m', processing_date) AS month,
        COALESCE(SUM(tax), 0) AS sales_tax
    FROM `{PROJECT_ID}.{DATASET_ID}.OrderDetails_raw`
    WHERE processing_date BETWEEN @start_date AND @end_date
        AND (voided IS NULL OR voided = 'false')
    GROUP BY month ORDER BY month
    """
    rows = client.query(q, job_config=_make_date_config(start, end)).result()
    return {r.month: float(r.sales_tax or 0) for r in rows}


def query_cash_undeposited(client: bigquery.Client, start: str, end: str) -> Dict[str, float]:
    """Monthly undeposited cash = Toast POS cash collected minus bank cash deposits."""
    # Cash collected at POS
    q_collected = f"""
    SELECT
        FORMAT_DATE('%Y-%m', processing_date) AS month,
        COALESCE(SUM(CASE WHEN payment_type = 'Cash' OR payment_type LIKE '%CASH%'
                     THEN total ELSE 0 END), 0) AS cash_collected
    FROM `{PROJECT_ID}.{DATASET_ID}.PaymentDetails_raw`
    WHERE processing_date BETWEEN @start_date AND @end_date
    GROUP BY month ORDER BY month
    """
    collected = {r.month: float(r.cash_collected or 0)
                 for r in client.query(q_collected, job_config=_make_date_config(start, end)).result()}

    # Cash deposited at bank — includes THREE patterns:
    #   1. Direct cash deposits ("cash deposit", "counter credit")
    #   2. Cash routed through Maurice's separate accounts (CHK 0227,
    #      CHK 9121) then transferred to operating. Per operator: these
    #      transfers-in typically fund payroll and originate from
    #      cash collected at LOV3.
    #   3. Plaid's TRANSFER_IN_ACCOUNT_TRANSFER category for the same flows.
    q_deposited = f"""
    SELECT
        FORMAT_DATE('%Y-%m', transaction_date) AS month,
        COALESCE(SUM(abs_amount), 0) AS cash_deposited
    FROM `{PROJECT_ID}.{DATASET_ID}.BankTransactions_raw`
    WHERE transaction_date BETWEEN @start_date AND @end_date
        AND transaction_type = 'credit'
        AND (
             -- Method 1: direct cash deposits
             LOWER(category) LIKE '%cash deposit%'
             OR LOWER(category) LIKE '%cash account transfer%'
             OR LOWER(description) LIKE '%counter credit%'
             -- Method 2: online-banking transfers in from Maurice's other
             -- LOV3 accounts (cash → reserve → operating). CHK 0227 and
             -- CHK 9121 are the two known intermediary accounts.
             OR LOWER(description) LIKE '%online banking transfer from chk 0227%'
             OR LOWER(description) LIKE '%online banking transfer from chk 9121%'
             -- Method 3: Plaid's transfer-in category for the same flows
             OR category = 'TRANSFER_IN_ACCOUNT_TRANSFER'
        )
    GROUP BY month ORDER BY month
    """
    deposited = {r.month: float(r.cash_deposited or 0)
                 for r in client.query(q_deposited, job_config=_make_date_config(start, end)).result()}

    # Undeposited = collected - deposited
    all_months = sorted(set(list(collected.keys()) + list(deposited.keys())))
    return {m: round(collected.get(m, 0) - deposited.get(m, 0), 2) for m in all_months}


# ── Data Assembly ────────────────────────────────────────────────────────────

def _sum_matching(cats: Dict[str, float], keywords: List[str]) -> float:
    """Sum expense categories whose name contains any of the keywords (case-insensitive)."""
    return sum(v for k, v in cats.items()
               if any(kw.lower() in k.lower() for kw in keywords))


def compute_pnl_for_month(
    rev: Dict, rev_cat: Dict, hookah_bank: float, hookah_pos: float,
    sales_tax: float, cash_undeposited: float, expenses: Dict[str, float]
) -> Dict[str, float]:
    """Compute all P&L line items for a single month.

    Revenue sourced from Toast POS (full picture):
      net_sales, tips, gratuity, sales_tax, cash_undeposited, hookah
    Expenses sourced from Bank of America (what hits the bank):
      debits by category
    """
    net_sales = rev.get("net_sales", 0)
    tips = rev.get("tips", 0)
    gratuity = rev.get("gratuity", 0)
    food_rev = rev_cat.get("food_rev", 0)
    liquor_rev = rev_cat.get("liquor_rev", 0)

    # Total hookah = POS (already in net_sales) + bank deposits (additive)
    hookah_total = round(hookah_pos + hookah_bank, 2)

    pass_through = round(tips + gratuity * GRAT_PASSTHROUGH_PCT, 2)
    # POS hookah is already in net_sales, so subtract it from other_rev to avoid double-count
    other_rev = round(max(net_sales - food_rev - liquor_rev - hookah_pos, 0), 2)
    # Total Operating Revenue: Toast POS full picture + bank hookah deposits
    # net_sales + tips + gratuity + cash_undeposited + sales_tax + hookah_bank
    total_revenue = round(
        net_sales + tips + gratuity + cash_undeposited + sales_tax + hookah_bank, 2
    )
    rev_denom = total_revenue if total_revenue > 0 else 1

    # COGS (includes Shisha COGS)
    food_cogs = _sum_matching(expenses, ["food cogs"])
    liquor_cogs = _sum_matching(expenses, ["liquor cogs", "shisha cogs"])
    supplies_cogs = _sum_matching(expenses, ["supplies & equipment", "supplies & smallwares"])
    total_cogs = round(food_cogs + liquor_cogs + supplies_cogs, 2)

    gross_profit = round(total_revenue - total_cogs, 2)

    # Labor (gross — includes tip pass-through, bonuses per business plan methodology)
    labor_gross = _sum_matching(expenses, ["3. labor", "labor cost", "payroll",
                                            "tip pass-through", "employee bonus"])

    # Marketing (includes PMG Artist, PPV)
    mkt_entertainment = _sum_matching(expenses, ["entertainment"])
    mkt_promoter = _sum_matching(expenses, ["promoter"])
    mkt_social = _sum_matching(expenses, ["social media"])
    mkt_flyers = _sum_matching(expenses, ["flyer", "digital ads", "print", "event flyer"])
    mkt_event = _sum_matching(expenses, ["event expense"])
    mkt_artist = _sum_matching(expenses, ["pmg artist", "artist booking"])
    mkt_ppv = _sum_matching(expenses, ["pay-per-view"])
    total_marketing = round(
        mkt_entertainment + mkt_promoter + mkt_social + mkt_flyers
        + mkt_event + mkt_artist + mkt_ppv, 2
    )

    # OPEX (includes uniforms, legal fees, chargebacks)
    opex_rent = _sum_matching(expenses, ["rent", "cam", "property tax"])
    opex_taxes = _sum_matching(expenses, ["5. operating expenses (opex)/taxes"])
    opex_security = _sum_matching(expenses, ["security"])
    opex_insurance = _sum_matching(expenses, ["insurance"])
    opex_bussers = _sum_matching(expenses, ["bussers & cleaners"])
    opex_contract_labor = _sum_matching(expenses, ["contract labor"])
    opex_cleaning = _sum_matching(expenses, ["janitorial services", "cleaning", "janitorial"])
    opex_utilities = _sum_matching(expenses, ["electric", "gas", "energy"])
    opex_pos_tech = _sum_matching(expenses, ["pos", "technology fee"])
    opex_software = _sum_matching(expenses, ["software", "subscription"])
    opex_phone = _sum_matching(expenses, ["phone", "internet"])
    opex_professional = _sum_matching(expenses, ["professional service", "legal", "accounting",
                                                   "consulting"])
    opex_permits = _sum_matching(expenses, ["permit", "license"])
    opex_bank_fees = _sum_matching(expenses, ["bank fee", "service charge"])
    opex_penalties = _sum_matching(expenses, ["penalty", "fine", "late fee"])
    opex_admin = _sum_matching(expenses, ["admin & office", "uniform"])
    opex_lighting = _sum_matching(expenses, ["lighting", "sound", "av"])
    opex_other = _sum_matching(expenses, ["chargeback", "adjustment",
                                           "other income/expense",
                                           "other/uncategorized", "uncategorized"])
    total_opex = round(
        opex_rent + opex_taxes + opex_security + opex_insurance + opex_bussers
        + opex_contract_labor + opex_cleaning + opex_utilities + opex_pos_tech
        + opex_software + opex_phone + opex_professional + opex_permits
        + opex_bank_fees + opex_penalties + opex_admin + opex_lighting
        + opex_other, 2
    )

    total_expenses_operating = round(total_cogs + labor_gross + total_marketing + total_opex, 2)

    # Facility / CapEx
    cap_construction = _sum_matching(expenses, ["construction", "build out"])
    cap_equipment = _sum_matching(expenses, ["capital equipment"])
    cap_repairs = _sum_matching(expenses, ["repair", "maintenance"])
    total_capex = round(cap_construction + cap_equipment + cap_repairs, 2)

    # G&A / Owner's Compensation (includes internal account transfers)
    ga_owner_draws = _sum_matching(expenses, ["owner draws"])
    ga_discretionary = _sum_matching(expenses, ["owner discretionary"])
    ga_meals = _sum_matching(expenses, ["personal meals"])
    ga_transportation = _sum_matching(expenses, ["6. general & administrative / corporate/transportation"])
    ga_travel = _sum_matching(expenses, ["owner travel", "travel & entertainment", "travel & lodging"])
    ga_competitive = _sum_matching(expenses, ["competitive research"])
    ga_credit_card = _sum_matching(expenses, ["credit card payments"])
    ga_other = _sum_matching(expenses, ["equity injection", "non-transaction",
                                         "operating account credit", "cash withdrawal",
                                         "internal account transfer"])
    total_ga = round(
        ga_owner_draws + ga_discretionary + ga_meals + ga_transportation
        + ga_travel + ga_competitive + ga_credit_card + ga_other, 2
    )

    total_all_expenses = round(total_expenses_operating + total_ga + total_capex, 2)
    ebitda = round(total_revenue - total_all_expenses, 2)

    # Memo: pass-through amount for disclosure note
    pass_through_memo = pass_through

    return {
        "food_rev": food_rev,
        "liquor_rev": liquor_rev,
        "hookah_rev": hookah_total,
        "other_rev": other_rev,
        "gratuity": gratuity,
        "tips": tips,
        "cash_undeposited": cash_undeposited,
        "sales_tax": sales_tax,
        "total_revenue": total_revenue,
        "food_cogs": food_cogs,
        "liquor_cogs": liquor_cogs,
        "supplies_cogs": supplies_cogs,
        "total_cogs": total_cogs,
        "gross_profit": gross_profit,
        "gross_margin_pct": round(gross_profit / rev_denom, 4),
        "labor_gross": labor_gross,
        "labor_pct": round(labor_gross / rev_denom, 4),
        "mkt_entertainment": mkt_entertainment,
        "mkt_promoter": mkt_promoter,
        "mkt_artist": mkt_artist,
        "mkt_social": mkt_social,
        "mkt_flyers": mkt_flyers,
        "mkt_event": mkt_event,
        "mkt_ppv": mkt_ppv,
        "total_marketing": total_marketing,
        "opex_rent": opex_rent,
        "opex_taxes": opex_taxes,
        "opex_security": opex_security,
        "opex_insurance": opex_insurance,
        "opex_bussers": opex_bussers,
        "opex_contract_labor": opex_contract_labor,
        "opex_cleaning": opex_cleaning,
        "opex_utilities": opex_utilities,
        "opex_pos_tech": opex_pos_tech,
        "opex_software": opex_software,
        "opex_phone": opex_phone,
        "opex_professional": opex_professional,
        "opex_permits": opex_permits,
        "opex_bank_fees": opex_bank_fees,
        "opex_penalties": opex_penalties,
        "opex_admin": opex_admin,
        "opex_lighting": opex_lighting,
        "opex_other": opex_other,
        "total_opex": total_opex,
        "total_expenses_operating": total_expenses_operating,
        "cap_construction": cap_construction,
        "cap_equipment": cap_equipment,
        "cap_repairs": cap_repairs,
        "total_capex": total_capex,
        "ga_owner_draws": ga_owner_draws,
        "ga_discretionary": ga_discretionary,
        "ga_meals": ga_meals,
        "ga_transportation": ga_transportation,
        "ga_travel": ga_travel,
        "ga_competitive": ga_competitive,
        "ga_credit_card": ga_credit_card,
        "ga_other": ga_other,
        "total_ga": total_ga,
        "total_all_expenses": total_all_expenses,
        "ebitda": ebitda,
        "ebitda_pct": round(ebitda / rev_denom, 4),
        "pass_through_memo": pass_through_memo,
    }


def sum_monthly_data(monthly: Dict[str, Dict[str, float]]) -> Dict[str, float]:
    """Sum all monthly P&L dicts into a YTD total."""
    ytd: Dict[str, float] = {}
    for month_data in monthly.values():
        for key, val in month_data.items():
            ytd[key] = ytd.get(key, 0) + val
    # Recompute percentages from YTD totals
    rev = ytd.get("total_revenue", 0) or 1
    ytd["gross_margin_pct"] = round(ytd.get("gross_profit", 0) / rev, 4)
    ytd["labor_pct"] = round(ytd.get("labor_gross", 0) / rev, 4)
    ytd["ebitda_pct"] = round(ytd.get("ebitda", 0) / rev, 4)
    return ytd


def query_period(client: bigquery.Client, start: str, end: str) -> Tuple[List[str], Dict[str, Dict], Dict[str, float]]:
    """Run all queries for a period and return (months, monthly_data, ytd_data).

    Revenue: Toast POS (full picture) + bank hookah deposits
    Expenses: Bank of America debits (CC transactions + limited cash deposits)
    """
    log.info(f"Querying period {start} to {end}...")

    rev = query_monthly_revenue(client, start, end)
    rev_cat = query_revenue_by_category(client, start, end)
    hookah_bank = query_hookah_revenue_bank(client, start, end)
    hookah_pos = query_hookah_revenue_pos(client, start, end)
    sales_tax = query_sales_tax(client, start, end)
    cash_undeposited = query_cash_undeposited(client, start, end)
    expenses = query_expenses_by_category(client, start, end)

    # Apply hookah reclass. Compare on YYYY-MM only — the previous
    # start <= m <= end check silently dropped the first month of any
    # period that started on the 1st (e.g., "2026-04-01" > "2026-04" in
    # lexicographic order, so April was skipped in the standalone Q2 P&L).
    for m, amt in HOOKAH_RECLASS.items():
        if start[:7] <= m <= end[:7]:
            hookah_bank[m] = hookah_bank.get(m, 0) + amt

    all_months = sorted(set(
        list(rev.keys()) + list(rev_cat.keys())
        + list(hookah_bank.keys()) + list(hookah_pos.keys())
        + list(sales_tax.keys()) + list(cash_undeposited.keys())
        + list(expenses.keys())
    ))

    monthly_data: Dict[str, Dict[str, float]] = {}
    for m in all_months:
        monthly_data[m] = compute_pnl_for_month(
            rev.get(m, {}),
            rev_cat.get(m, {}),
            hookah_bank.get(m, 0),
            hookah_pos.get(m, 0),
            sales_tax.get(m, 0),
            cash_undeposited.get(m, 0),
            expenses.get(m, {}),
        )

    ytd = sum_monthly_data(monthly_data)
    log.info(f"  {len(all_months)} months loaded, YTD revenue: ${ytd.get('total_revenue', 0):,.0f}")
    return all_months, monthly_data, ytd


# ── Excel Writer ─────────────────────────────────────────────────────────────

def _month_label(m: str) -> str:
    """Convert '2025-01' to 'Jan 2025'."""
    parts = m.split("-")
    return f"{calendar.month_abbr[int(parts[1])]} {parts[0]}"


def write_pnl_sheet(
    wb: Workbook,
    sheet_name: str,
    period_label: str,
    months: List[str],
    monthly_data: Dict[str, Dict[str, float]],
    ytd_data: Dict[str, float],
):
    """Write one P&L sheet to the workbook."""
    ws = wb.create_sheet(title=sheet_name)
    num_months = len(months)
    ytd_col = num_months + 2  # col B..B+n-1 = months, then YTD
    pct_col = ytd_col + 1     # % of Revenue
    last_col = pct_col

    # Column widths
    ws.column_dimensions["A"].width = 42
    for i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # ── Header block ──
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value="LOV3|HTX")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center")

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value="Income Statement (Profit & Loss)")
    c.font = Font(name="Calibri", size=12, bold=True)
    c.alignment = Alignment(horizontal="center")

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=period_label)
    c.font = SUBTITLE_FONT
    c.alignment = Alignment(horizontal="center")

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=f"Prepared: {date.today().strftime('%B %d, %Y')}")
    c.font = Font(name="Calibri", size=9, italic=True, color="666666")
    c.alignment = Alignment(horizontal="center")

    # ── Column headers ──
    row = 6
    ws.cell(row=row, column=1, value="").font = HEADER_FONT
    for i, m in enumerate(months):
        c = ws.cell(row=row, column=i + 2, value=_month_label(m))
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=row, column=ytd_col, value="YTD Total")
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=row, column=pct_col, value="% of Rev")
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")

    # ── P&L rows ──
    row = 7
    for label, key, indent, line_type in PNL_STRUCTURE:
        if line_type == "blank":
            row += 1
            continue

        # Label cell
        display_label = ("  " * indent + label) if indent else label
        c = ws.cell(row=row, column=1, value=display_label)

        if line_type == "section":
            c.font = SECTION_FONT
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).fill = SECTION_FILL
            row += 1
            continue

        if line_type == "memo":
            # Memo note with pass-through amount in YTD column
            c.font = Font(name="Calibri", size=9, italic=True, color="666666")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ytd_col - 1)
            if key:
                ytd_val = ytd_data.get(key, 0)
                cell = ws.cell(row=row, column=ytd_col, value=round(ytd_val, 2))
                cell.number_format = CURRENCY_FMT
                cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
            row += 1
            continue

        # Data cells
        if key:
            is_pct = line_type == "pct_line"
            is_credit = line_type == "credit"

            for i, m in enumerate(months):
                val = monthly_data.get(m, {}).get(key, 0)
                cell = ws.cell(row=row, column=i + 2)
                if is_pct:
                    cell.value = val
                    cell.number_format = PCT_FMT
                elif is_credit and val > 0:
                    cell.value = -val
                    cell.number_format = CURRENCY_FMT
                    cell.font = RED_FONT
                else:
                    cell.value = round(val, 2)
                    cell.number_format = CURRENCY_FMT
                    if val < 0:
                        cell.font = RED_FONT

            # YTD
            ytd_val = ytd_data.get(key, 0)
            cell = ws.cell(row=row, column=ytd_col)
            if is_pct:
                cell.value = ytd_val
                cell.number_format = PCT_FMT
            elif is_credit and ytd_val > 0:
                cell.value = -ytd_val
                cell.number_format = CURRENCY_FMT
                cell.font = RED_FONT
            else:
                cell.value = round(ytd_val, 2)
                cell.number_format = CURRENCY_FMT
                if ytd_val < 0:
                    cell.font = RED_FONT

            # % of Revenue
            if not is_pct:
                rev = ytd_data.get("total_revenue", 0) or 1
                pct = ytd_val / rev if not is_credit else -ytd_val / rev
                cell = ws.cell(row=row, column=pct_col)
                cell.value = abs(pct) if is_credit else pct
                cell.number_format = PCT_FMT

        # Formatting by line type
        if line_type in ("total", "subtotal"):
            c.font = TOTAL_FONT
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).border = THIN_TOP
            c.font = TOTAL_FONT
        elif line_type == "net_income":
            c.font = Font(name="Calibri", size=12, bold=True)
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).border = DOUBLE_BOTTOM
                ws.cell(row=row, column=col).font = Font(name="Calibri", size=11, bold=True)
            c.font = Font(name="Calibri", size=12, bold=True)

        row += 1

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.print_area = f"A1:{get_column_letter(last_col)}{row}"
    log.info(f"  Sheet '{sheet_name}' written ({row} rows)")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    client = bigquery.Client(project=PROJECT_ID)

    # Period 1: 2025 Year-End
    months_2025, data_2025, ytd_2025 = query_period(client, "2025-01-01", "2025-12-31")

    # Period 2: Jan–Mar 2026 Interim
    months_2026, data_2026, ytd_2026 = query_period(client, "2026-01-01", "2026-03-31")

    # Period 3: Apr–Jun 2026 Q2 Interim
    months_q2_2026, data_q2_2026, ytd_q2_2026 = query_period(
        client, "2026-04-01", "2026-06-30")

    # Period 4: H1 2026 YTD (Jan–Jun 2026) — lender wants YTD view too
    months_h1_2026, data_h1_2026, ytd_h1_2026 = query_period(
        client, "2026-01-01", "2026-06-30")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_pnl_sheet(
        wb, "2025 Year-End P&L",
        "For the Year Ended December 31, 2025",
        months_2025, data_2025, ytd_2025,
    )
    write_pnl_sheet(
        wb, "Q1 2026 P&L",
        "For the Three Months Ended March 31, 2026",
        months_2026, data_2026, ytd_2026,
    )
    write_pnl_sheet(
        wb, "Q2 2026 P&L",
        "For the Three Months Ended June 30, 2026",
        months_q2_2026, data_q2_2026, ytd_q2_2026,
    )
    write_pnl_sheet(
        wb, "H1 2026 YTD P&L",
        "For the Six Months Ended June 30, 2026",
        months_h1_2026, data_h1_2026, ytd_h1_2026,
    )

    filename = "LOV3_HTX_Financial_Statements_SBA.xlsx"
    wb.save(filename)
    log.info(f"Saved: {filename}")

    # Standalone Q1 2026 file (kept for continuity)
    wb2 = Workbook()
    wb2.remove(wb2.active)
    write_pnl_sheet(
        wb2, "Q1 2026 P&L",
        "For the Three Months Ended March 31, 2026",
        months_2026, data_2026, ytd_2026,
    )
    q1_filename = "LOV3_HTX_Q1_2026_PL.xlsx"
    wb2.save(q1_filename)
    log.info(f"Saved: {q1_filename}")

    # Standalone Q2 2026 file
    wb3 = Workbook()
    wb3.remove(wb3.active)
    write_pnl_sheet(
        wb3, "Q2 2026 P&L",
        "For the Three Months Ended June 30, 2026",
        months_q2_2026, data_q2_2026, ytd_q2_2026,
    )
    q2_filename = "LOV3_HTX_Q2_2026_PL.xlsx"
    wb3.save(q2_filename)
    log.info(f"Saved: {q2_filename}")

    # Standalone H1 2026 YTD file (for lender's YTD view)
    wb4 = Workbook()
    wb4.remove(wb4.active)
    write_pnl_sheet(
        wb4, "H1 2026 YTD P&L",
        "For the Six Months Ended June 30, 2026",
        months_h1_2026, data_h1_2026, ytd_h1_2026,
    )
    h1_filename = "LOV3_HTX_H1_2026_YTD_PL.xlsx"
    wb4.save(h1_filename)
    log.info(f"Saved: {h1_filename}")

    print(f"\nDone! Files saved:")
    print(f"  {filename}")
    print(f"  {q1_filename}")
    print(f"  {q2_filename}")
    print(f"  {h1_filename}")


if __name__ == "__main__":
    main()
